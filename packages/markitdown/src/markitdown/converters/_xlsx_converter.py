import sys
import re
from typing import BinaryIO, Any
from ._html_converter import HtmlConverter
from .._base_converter import DocumentConverter, DocumentConverterResult
from .._exceptions import MissingDependencyException, MISSING_DEPENDENCY_MESSAGE
from .._stream_info import StreamInfo

try:
    from paddleocr import PaddleOCR
    from PIL import Image
    import io
    _has_ocr_deps = True
except ImportError:
    import io
    _has_ocr_deps = False

# Try loading optional (but in this case, required) dependencies
# Save reporting of any exceptions for later
_xlsx_dependency_exc_info = None
try:
    import pandas as pd
    import openpyxl
except ImportError:
    _xlsx_dependency_exc_info = sys.exc_info()

_xls_dependency_exc_info = None
try:
    import pandas as pd  # noqa: F811
    import xlrd
except ImportError:
    _xls_dependency_exc_info = sys.exc_info()

ACCEPTED_XLSX_MIME_TYPE_PREFIXES = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
]
ACCEPTED_XLSX_FILE_EXTENSIONS = [".xlsx"]

ACCEPTED_XLS_MIME_TYPE_PREFIXES = [
    "application/vnd.ms-excel",
    "application/excel",
]
ACCEPTED_XLS_FILE_EXTENSIONS = [".xls"]


class XlsxConverter(DocumentConverter):
    """
    Converts XLSX files to Markdown, with each sheet presented as a separate Markdown table.
    """

    def __init__(self):
        super().__init__()
        self._html_converter = HtmlConverter()

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()

        if extension in ACCEPTED_XLSX_FILE_EXTENSIONS:
            return True

        for prefix in ACCEPTED_XLSX_MIME_TYPE_PREFIXES:
            if mimetype.startswith(prefix):
                return True

        return False

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> DocumentConverterResult:
        # Check the dependencies
        if _xlsx_dependency_exc_info is not None:
            raise MissingDependencyException(
                MISSING_DEPENDENCY_MESSAGE.format(
                    converter=type(self).__name__, extension=".xlsx", feature="xlsx"
                )
            ) from _xlsx_dependency_exc_info[1].with_traceback(_xlsx_dependency_exc_info[2])

        # Feature flags
        detect_hierarchy = kwargs.get("detect_hierarchy", True)
        
        # Load workbook with openpyxl for structural metadata and image extraction
        file_stream.seek(0)
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        ocr = None
        
        md_content = ""
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            md_content += f"# {sheet_name}\n\n"
            
            # --- 1. OCR Pre-processing ---
            ocr_map = {}
            if _has_ocr_deps and hasattr(ws, "_images") and ws._images:
                if ocr is None:
                    ocr = PaddleOCR(use_angle_cls=True, lang='fr', show_log=False)
                for img in ws._images:
                    try:
                        anchor = img.anchor
                        img_row, img_col = anchor._from.row + 1, anchor._from.col + 1
                        img_data = img.ref.getvalue()
                        res = ocr.ocr(img_data, cls=True)
                        txt = " ".join([line[1][0] for line in res[0]]).strip() if res and res[0] else ""
                        if txt: ocr_map[(img_row, img_col)] = txt
                    except Exception: continue

            if detect_hierarchy:
                # --- 2. Structural/Surgical Reconstruction ---
                merge_lookup = {}
                for mr in ws.merged_cells.ranges:
                    for r in range(mr.min_row, mr.max_row + 1):
                        for c in range(mr.min_col, mr.max_col + 1):
                            merge_lookup[(r, c)] = mr

                sections = [] 
                for r in range(1, ws.max_row + 1):
                    new_path = []
                    for c_idx in [1, 2]: # Columns A and B
                        mr = merge_lookup.get((r, c_idx))
                        cell = ws.cell(row=mr.min_row if mr else r, column=mr.min_col if mr else c_idx)
                        val = str(cell.value or ocr_map.get((cell.row, cell.column), "")).strip()
                        if val and val.lower() not in ["none", "nan", "unnamed"]:
                            if not new_path or val != new_path[-1]:
                                new_path.append(val)
                    
                    if not sections or new_path != sections[-1]["path"]:
                        if new_path:
                            sections.append({"path": new_path, "rows": [r]})
                    elif sections:
                        sections[-1]["rows"].append(r)

                for section in sections:
                    if section["path"]:
                        md_content += f"### {' > '.join(section['path'])}\n\n"
                    
                    table_data = [] # (Label, Code, Value)
                    code_pattern = re.compile(r"^[A-ZØ]?[A-Z0-9][A-Z0-9]$")
                    
                    for r_idx in section["rows"]:
                        temp_label = None
                        for c_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=r_idx, column=c_idx)
                            val = str(cell.value or ocr_map.get((r_idx, c_idx), "")).strip()
                            if not val or val.lower() in ["none", "nan"]: continue
                            
                            if code_pattern.match(val):
                                if temp_label:
                                    val_cell = ws.cell(row=r_idx, column=c_idx + 1)
                                    amount = str(val_cell.value).strip() if val_cell.value is not None else ""
                                    table_data.append((temp_label, val, amount))
                                    temp_label = None
                            elif len(val) > 2:
                                temp_label = val
                    
                    if table_data:
                        md_content += "| Label | Code | Montant |\n"
                        md_content += "|---|---|---|\n"
                        for label, code, amount in table_data:
                            md_content += f"| {label} | {code} | {amount} |\n"
                        md_content += "\n"
            else:
                # --- 3. Standard Fallback Path ---
                file_stream.seek(0)
                df = pd.read_excel(file_stream, sheet_name=sheet_name, engine="openpyxl")
                md_content += self._html_converter.convert_string(df.to_html(index=False), **kwargs).markdown.strip() + "\n\n"

        return DocumentConverterResult(markdown=md_content.strip())


class XlsConverter(DocumentConverter):
    """
    Converts XLS files to Markdown, with each sheet presented as a separate Markdown table.
    """

    def __init__(self):
        super().__init__()
        self._html_converter = HtmlConverter()

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()

        if extension in ACCEPTED_XLS_FILE_EXTENSIONS:
            return True

        for prefix in ACCEPTED_XLS_MIME_TYPE_PREFIXES:
            if mimetype.startswith(prefix):
                return True

        return False

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> DocumentConverterResult:
        # Load the dependencies
        if _xls_dependency_exc_info is not None:
            raise MissingDependencyException(
                MISSING_DEPENDENCY_MESSAGE.format(
                    converter=type(self).__name__, extension=".xls", feature="xls"
                )
            ) from _xls_dependency_exc_info[1].with_traceback(_xls_dependency_exc_info[2])

        sheets = pd.read_excel(file_stream, sheet_name=None, engine="xlrd")
        md_content = ""
        for s in sheets:
            md_content += f"## {s}\n"
            html_content = sheets[s].to_html(index=False)
            md_content += (
                self._html_converter.convert_string(
                    html_content, **kwargs
                ).markdown.strip()
                + "\n\n"
            )

        return DocumentConverterResult(markdown=md_content.strip())
