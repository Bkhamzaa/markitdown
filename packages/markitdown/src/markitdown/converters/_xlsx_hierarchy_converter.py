import sys
import re
from typing import BinaryIO, Any, List, Tuple, Dict, Set, Optional
from ._html_converter import HtmlConverter
from .._base_converter import DocumentConverter, DocumentConverterResult
from .._exceptions import MissingDependencyException, MISSING_DEPENDENCY_MESSAGE
from .._stream_info import StreamInfo

try:
    from paddleocr import PaddleOCR
    from PIL import Image
    import io
    _has_ocr_deps = True
except ImportError:
    import io
    _has_ocr_deps = False

_xlsx_dependency_exc_info = None
try:
    import pandas as pd
    import openpyxl
except ImportError:
    _xlsx_dependency_exc_info = sys.exc_info()

_xls_dependency_exc_info = None
try:
    import pandas as pd  # noqa: F811
    import xlrd
except ImportError:
    _xls_dependency_exc_info = sys.exc_info()

ACCEPTED_XLSX_MIME_TYPE_PREFIXES = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
]
ACCEPTED_XLSX_FILE_EXTENSIONS = [".xlsx"]
ACCEPTED_XLS_MIME_TYPE_PREFIXES = [
    "application/vnd.ms-excel",
    "application/excel",
]
ACCEPTED_XLS_FILE_EXTENSIONS = [".xls"]


class XlsxConverter(DocumentConverter):
    """
    Universal Structural Discovery Engine for DGFIP-style forms.

    Discovers hierarchy purely from vertical merge patterns:
      - Vertical merges (row_span > 1, col_span <= 4) with label text -> section headers
      - Header depth derived from column position within column groups
      - Parallel panels (side-by-side sections) detected and emitted sequentially
      - Wide single-row merges without codes -> banners (## titles)
      - Wide merges starting with ( or * -> footnotes (blockquotes at the end)
    """

    def __init__(self):
        super().__init__()
        self._html_converter = HtmlConverter()

    def accepts(
        self, file_stream: BinaryIO, stream_info: StreamInfo, **kwargs: Any
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()
        return extension in ACCEPTED_XLSX_FILE_EXTENSIONS or any(
            mimetype.startswith(p) for p in ACCEPTED_XLSX_MIME_TYPE_PREFIXES
        )

    @staticmethod
    def _get_val(ws, r: int, c: int, merge_lookup: dict, ocr_map: dict) -> str:
        """Return the visible text for any cell, resolving merges and OCR."""
        if r < 1 or c < 1:
            return ""
        mr = merge_lookup.get((r, c))
        tr, tc = (mr.min_row, mr.min_col) if mr else (r, c)
        val = ws.cell(row=tr, column=tc).value
        # Fallback to OCR map if direct cell is empty
        if val is None:
            val = ocr_map.get((tr, tc), "")
        return str(val).strip() if val is not None else ""

    _CODE_RE = re.compile(r"^[A-Z0-9Ø][A-Z0-9]$")

    @classmethod
    def _is_code(cls, text: str) -> bool:
        return bool(cls._CODE_RE.match(text))

    @classmethod
    def _is_label(cls, text: str) -> bool:
        """Real label text (not a code, formula, or number)."""
        if not text or len(text) <= 2:
            return False
        if cls._is_code(text):
            return False
        if text.startswith("="):
            return False
        try:
            float(text.replace(",", ".").replace(" ", "").replace("\xa0", ""))
            return False
        except ValueError:
            return True

    @staticmethod
    def _is_note(text: str) -> bool:
        """Parenthetical / explanatory note — never use as primary label."""
        t = text.lower().strip()
        return (
            t.startswith("dont ")
            or t.startswith("dont\n")
            or t.startswith("précisez")
            or t.startswith("(")
            or t.startswith("si oui")
            or t.startswith("si non")
        )

    @staticmethod
    def _is_numeric(text: str) -> bool:
        if not text:
            return False
        try:
            float(
                text.replace(",", ".").replace(" ", "").replace("\xa0", "")
            )
            return True
        except (ValueError, AttributeError):
            return False

    @staticmethod
    def _render_table(data: List[Tuple[str, str, str]]) -> str:
        if not data:
            return ""
        out = "| Label | Code | Montant |\n|---|---|---|\n"
        for label, code, montant in data:
            out += f"| {label} | {code} | {montant} |\n"
        return out + "\n"

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,
    ) -> DocumentConverterResult:
        if _xlsx_dependency_exc_info is not None:
            raise MissingDependencyException(
                MISSING_DEPENDENCY_MESSAGE.format(
                    converter=type(self).__name__,
                    extension=".xlsx",
                    feature="xlsx",
                )
            )

        file_stream.seek(0)
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        ocr = None
        md = ""

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            md += f"# {sheet_name}\n\n"
            max_r = ws.max_row or 0
            max_c = ws.max_column or 0
            if max_r == 0 or max_c == 0:
                continue

            # --- OCR PRE-SCAN ---
            ocr_map: Dict[Tuple[int, int], str] = {}
            if _has_ocr_deps and hasattr(ws, "_images") and ws._images:
                if ocr is None:
                    ocr = PaddleOCR(use_angle_cls=True, lang="fr", show_log=False)
                for img_obj in ws._images:
                    try:
                        anchor = img_obj.anchor
                        ir, ic = anchor._from.row + 1, anchor._from.col + 1
                        res = ocr.ocr(img_obj.ref.getvalue(), cls=True)
                        txt = " ".join(l[1][0] for l in res[0]).strip() if res and res[0] else ""
                        if txt: ocr_map[(ir, ic)] = txt
                    except Exception: continue

            # --- MERGE MAP ---
            merge_lookup: Dict[Tuple[int, int], Any] = {}
            for mr in ws.merged_cells.ranges:
                for r in range(mr.min_row, mr.max_row + 1):
                    for c in range(mr.min_col, mr.max_col + 1):
                        merge_lookup[(r, c)] = mr

            gv = lambda r, c: self._get_val(ws, r, c, merge_lookup, ocr_map)

            rows_with_codes: Set[int] = set()
            for r in range(1, max_r + 1):
                for c in range(1, max_c + 1):
                    if self._is_code(gv(r, c)):
                        rows_with_codes.add(r)
                        break

            # --- CLASSIFY MERGES ---
            hierarchy_merges = []
            banner_rows = {}
            footnote_rows = set()

            for mr in ws.merged_cells.ranges:
                cspan = mr.max_col - mr.min_col + 1
                rspan = mr.max_row - mr.min_row + 1
                wide = cspan / max_c >= 0.60

                if wide and rspan == 1:
                    txt = gv(mr.min_row, mr.min_col)
                    if not txt: continue
                    if txt.startswith("(") or txt.startswith("*"): footnote_rows.add(mr.min_row)
                    elif mr.min_row not in rows_with_codes: banner_rows[mr.min_row] = (mr, txt)
                    continue
                if wide and rspan > 1:
                    for rr in range(mr.min_row, mr.max_row + 1): footnote_rows.add(rr)
                    continue

                if rspan > 1 and cspan <= 4:
                    txt = gv(mr.min_row, mr.min_col)
                    if not txt: # OCR bounding box fallback
                        for (or_, oc_), otxt in ocr_map.items():
                            if mr.min_row <= or_ <= mr.max_row and mr.min_col <= oc_ <= mr.max_col:
                                txt = otxt; break
                    if self._is_label(txt): hierarchy_merges.append(mr)

            # --- DEPTH & REGIONS ---
            h_cols = sorted(set(m.min_col for m in hierarchy_merges))
            col_groups = []
            if h_cols:
                grp = [h_cols[0]]
                for c in h_cols[1:]:
                    if c - grp[-1] >= 3:
                        col_groups.append(grp); grp = [c]
                    else: grp.append(c)
                col_groups.append(grp)

            BASE_DEPTH = 3
            col_depth = {}
            for grp in col_groups:
                for idx, c in enumerate(sorted(grp)): col_depth[c] = BASE_DEPTH + idx

            h_by_start = {}
            for mr in hierarchy_merges: h_by_start.setdefault(mr.min_row, []).append(mr)
            for k in h_by_start: h_by_start[k].sort(key=lambda m: col_depth.get(m.min_col, 99))

            parallel_rows = {}
            for row_key, merges in h_by_start.items():
                if len(merges) < 2: continue
                hits = {}
                for mr in merges:
                    for gi, grp in enumerate(col_groups):
                        if mr.min_col in grp: hits.setdefault(gi, mr)
                if len(hits) >= 2:
                    panels = []
                    for gi in sorted(hits):
                        p_mr = hits[gi]
                        c_lo, c_hi = min(col_groups[gi]), (min(col_groups[gi+1])-1 if gi+1 < len(col_groups) else max_c)
                        panels.append((p_mr, c_lo, c_hi))
                    parallel_rows[row_key] = panels

            # --- ROW WALK ---
            emitted = set()
            h_stack = []
            table_buf = []
            processed = set()

            def flush():
                nonlocal table_buf, md
                if table_buf: md += self._render_table(table_buf); table_buf = []

            def emit_hdr(mr):
                nonlocal md
                txt = gv(mr.min_row, mr.min_col)
                if not txt: # Fallback
                    for (or_, oc_), otxt in ocr_map.items():
                        if mr.min_row <= or_ <= mr.max_row and mr.min_col <= oc_ <= mr.max_col:
                            txt = otxt; break
                if not txt or txt in emitted: return
                flush()
                d = col_depth.get(mr.min_col, BASE_DEPTH)
                md += f"{'#' * d} {txt}\n\n"
                emitted.add(txt)

            def extract_row(r, c_lo, c_hi):
                nonlocal table_buf
                for c in range(c_lo, c_hi + 1):
                    v = gv(r, c)
                    if not self._is_code(v): continue
                    label = ""
                    for lc in range(c - 1, c_lo - 1, -1):
                        lv = gv(r, lc)
                        if self._is_label(lv) and not self._is_note(lv):
                            if not any(lv == h[2] for h in h_stack):
                                label = lv; break
                    montant = ""
                    for rc in range(c + 1, min(c + 3, c_hi + 1)):
                        rv = gv(r, rc)
                        if rv and self._is_numeric(rv): montant = rv; break
                    entry = (label, v, montant)
                    if entry not in table_buf: table_buf.append(entry)

            for r in range(1, max_r + 1):
                if r in processed or r in footnote_rows: continue
                if r in banner_rows:
                    _, txt = banner_rows[r]
                    if txt and txt not in emitted: flush(); md += f"## {txt}\n\n"; emitted.add(txt)
                    continue
                h_stack = [h for h in h_stack if r <= h[0].max_row]
                if r in parallel_rows:
                    flush()
                    for p_mr, clo, chi in parallel_rows[r]:
                        emit_hdr(p_mr)
                        for pr in range(p_mr.min_row, p_mr.max_row + 1):
                            extract_row(pr, clo, chi); processed.add(pr)
                        flush()
                    continue
                if r in h_by_start:
                    for mr in h_by_start[r]:
                        txt = gv(mr.min_row, mr.min_col)
                        if not txt:
                            for (or_, oc_), otxt in ocr_map.items():
                                if mr.min_row<=or_<=mr.max_row and mr.min_col<=oc_<=mr.max_col:
                                    txt=otxt; break
                        if txt and txt not in emitted:
                            emit_hdr(mr); d = col_depth.get(mr.min_col, BASE_DEPTH); h_stack.append((mr, d, txt))
                extract_row(r, 1, max_c)

            flush()
            fn_done = set()
            for r in sorted(footnote_rows):
                mr_fn = merge_lookup.get((r, 1))
                if mr_fn and r == mr_fn.min_row:
                    txt = gv(r, 1)
                    if txt and txt not in fn_done and txt not in emitted:
                        md += f"\n> {txt}\n"; fn_done.add(txt)
            md += "\n"

        return DocumentConverterResult(markdown=md.strip())


class XlsConverter(DocumentConverter):
    def __init__(self):
        super().__init__()
        self._html_converter = HtmlConverter()
    def accepts(self, file_stream, stream_info, **kwargs):
        mimetype, extension = (stream_info.mimetype or "").lower(), (stream_info.extension or "").lower()
        return extension in ACCEPTED_XLS_FILE_EXTENSIONS or any(mimetype.startswith(p) for p in ACCEPTED_XLS_MIME_TYPE_PREFIXES)
    def convert(self, file_stream, stream_info, **kwargs):
        if _xls_dependency_exc_info: raise MissingDependencyException(MISSING_DEPENDENCY_MESSAGE.format(converter=type(self).__name__, extension=".xls", feature="xls"))
        sheets = pd.read_excel(file_stream, sheet_name=None, engine="xlrd")
        md = ""
        for s in sheets:
            md += f"## {s}\n"
            md += self._html_converter.convert_string(sheets[s].to_html(index=False), **kwargs).markdown.strip() + "\n\n"
        return DocumentConverterResult(markdown=md.strip())
